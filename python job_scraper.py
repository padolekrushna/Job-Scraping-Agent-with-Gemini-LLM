import os
import re
import time
import json
import requests
import pandas as pd
from datetime import datetime
from urllib.parse import urljoin, urlparse
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import google.generativeai as genai
from PyPDF2 import PdfReader
from docx import Document
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class JobScrapingAgent:
    def __init__(self, gemini_api_key):
        """Initialize the job scraping agent with Gemini API key"""
        self.gemini_api_key = gemini_api_key
        genai.configure(api_key=gemini_api_key)
        self.model = genai.GenerativeModel('gemini-pro')
        self.user_skills = []
        self.user_experience = ""
        self.jobs_data = []
        
    def extract_resume_info(self, resume_path):
        """Extract skills and experience from user's resume"""
        logger.info(f"Extracting information from resume: {resume_path}")
        
        if not os.path.exists(resume_path):
            raise FileNotFoundError(f"Resume file not found: {resume_path}")
            
        file_extension = os.path.splitext(resume_path)[1].lower()
        
        if file_extension == '.pdf':
            text = self._extract_pdf_text(resume_path)
        elif file_extension in ['.docx', '.doc']:
            text = self._extract_docx_text(resume_path)
        else:
            raise ValueError("Unsupported file format. Please provide PDF or DOCX file.")
            
        # Use Gemini to extract structured information
        prompt = f"""
        Analyze the following resume text and extract:
        1. A list of technical and professional skills (as a Python list)
        2. A summary of professional experience (as a string)
        
        Resume text:
        {text}
        
        Respond ONLY in JSON format with keys "skills" and "experience".
        """
        
        try:
            response = self.model.generate_content(prompt)
            # Clean the response to ensure valid JSON
            response_text = response.text.strip()
            if response_text.startswith("```json"):
                response_text = response_text[7:]
            if response_text.endswith("```"):
                response_text = response_text[:-3]
                
            resume_info = json.loads(response_text)
            self.user_skills = resume_info.get("skills", [])
            self.user_experience = resume_info.get("experience", "")
            
            logger.info(f"Extracted {len(self.user_skills)} skills from resume")
            return self.user_skills, self.user_experience
            
        except Exception as e:
            logger.error(f"Error extracting resume info: {str(e)}")
            # Fallback to basic extraction
            self.user_skills = self._basic_skill_extraction(text)
            self.user_experience = self._basic_experience_extraction(text)
            return self.user_skills, self.user_experience
    
    def _extract_pdf_text(self, pdf_path):
        """Extract text from PDF file"""
        reader = PdfReader(pdf_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text
    
    def _extract_docx_text(self, docx_path):
        """Extract text from DOCX file"""
        doc = Document(docx_path)
        text = ""
        for para in doc.paragraphs:
            text += para.text + "\n"
        return text
    
    def _basic_skill_extraction(self, text):
        """Basic skill extraction using keyword matching"""
        # Common tech skill keywords
        tech_keywords = [
            'python', 'java', 'javascript', 'c++', 'c#', 'sql', 'html', 'css', 
            'react', 'angular', 'vue', 'node.js', 'django', 'flask', 'spring', 
            'aws', 'azure', 'gcp', 'docker', 'kubernetes', 'jenkins', 'git',
            'machine learning', 'deep learning', 'nlp', 'computer vision',
            'data analysis', 'data science', 'big data', 'hadoop', 'spark',
            'tableau', 'power bi', 'excel', 'r', 'matlab', 'scala', 'go', 'rust'
        ]
        
        found_skills = []
        text_lower = text.lower()
        for keyword in tech_keywords:
            if keyword in text_lower:
                found_skills.append(keyword.title())
                
        return list(set(found_skills))
    
    def _basic_experience_extraction(self, text):
        """Basic experience extraction"""
        # Look for experience sections
        lines = text.split('\n')
        experience_lines = []
        in_experience_section = False
        
        for line in lines:
            if any(word in line.lower() for word in ['experience', 'work history', 'employment']):
                in_experience_section = True
                continue
            if in_experience_section and any(word in line.lower() for word in ['education', 'skills', 'certifications']):
                break
            if in_experience_section:
                experience_lines.append(line)
                
        return '\n'.join(experience_lines) if experience_lines else "Professional experience details not found."
    
    def scrape_jobs(self, job_title, location, num_pages=3):
        """
        Scrape job listings from multiple sources
        Currently supports: LinkedIn, Indeed (via scraping - note: check robots.txt)
        """
        logger.info(f"Scraping jobs for '{job_title}' in '{location}'")
        
        # We'll use Selenium for dynamic content
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)
        
        try:
            # Scrape from LinkedIn (simplified - note: LinkedIn has anti-scraping measures)
            self._scrape_linkedin_jobs(driver, job_title, location, num_pages)
            
            # Scrape from Indeed
            self._scrape_indeed_jobs(driver, job_title, location, num_pages)
            
        finally:
            driver.quit()
            
        logger.info(f"Scraped {len(self.jobs_data)} job listings")
        return self.jobs_data
    
    def _scrape_linkedin_jobs(self, driver, job_title, location, num_pages):
        """Scrape job listings from LinkedIn"""
        # Note: LinkedIn has strict anti-scraping measures
        # This is a simplified approach that may not work reliably
        base_url = "https://www.linkedin.com/jobs/search"
        params = {
            "keywords": job_title,
            "location": location,
            "f_TPR": "r86400",  # Past 24 hours
            "position": "1",
            "pageNum": "0"
        }
        
        # Build URL
        url = base_url + "?" + "&".join([f"{k}={v}" for k, v in params.items()])
        logger.info(f"Scraping LinkedIn: {url}")
        
        try:
            driver.get(url)
            time.sleep(3)
            
            for page in range(num_pages):
                # Scroll to load more jobs
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(2)
                
                soup = BeautifulSoup(driver.page_source, 'html.parser')
                job_cards = soup.find_all('div', class_='base-card')
                
                for card in job_cards:
                    try:
                        title_elem = card.find('h3', class_='base-search-card__title')
                        company_elem = card.find('h4', class_='base-search-card__subtitle')
                        link_elem = card.find('a', class_='base-card__full-link')
                        description_elem = card.find('p', class_='job-search-card__snippet')
                        
                        if title_elem and link_elem:
                            job = {
                                'title': title_elem.text.strip(),
                                'company': company_elem.text.strip() if company_elem else "N/A",
                                'link': link_elem['href'],
                                'description': description_elem.text.strip() if description_elem else "",
                                'source': 'LinkedIn'
                            }
                            self.jobs_data.append(job)
                    except Exception as e:
                        logger.warning(f"Error parsing LinkedIn job card: {str(e)}")
                        continue
                
                # Try to go to next page
                try:
                    next_button = driver.find_element("xpath", "//button[@aria-label='Next']")
                    if next_button.is_enabled():
                        next_button.click()
                        time.sleep(3)
                    else:
                        break
                except:
                    break
                    
        except Exception as e:
            logger.error(f"Error scraping LinkedIn: {str(e)}")
    
    def _scrape_indeed_jobs(self, driver, job_title, location, num_pages):
        """Scrape job listings from Indeed"""
        base_url = "https://www.indeed.com/jobs"
        params = {
            "q": job_title,
            "l": location,
            "fromage": "1",  # Last 24 hours
            "sort": "date"
        }
        
        # Build URL
        url = base_url + "?" + "&".join([f"{k}={v.replace(' ', '+')}" for k, v in params.items()])
        logger.info(f"Scraping Indeed: {url}")
        
        try:
            for page in range(num_pages):
                page_url = f"{url}&start={page * 10}"
                driver.get(page_url)
                time.sleep(2)
                
                soup = BeautifulSoup(driver.page_source, 'html.parser')
                job_cards = soup.find_all('div', class_='job_seen_beacon')
                
                for card in job_cards:
                    try:
                        title_elem = card.find('h2', class_='jobTitle')
                        company_elem = card.find('span', class_='companyName')
                        link_elem = card.find('a', class_='jcs-JobTitle')
                        description_elem = card.find('div', class_='job-snippet')
                        
                        if title_elem and link_elem:
                            # Extract title text
                            title = ""
                            for span in title_elem.find_all('span'):
                                if span.get('title'):
                                    title = span['title']
                                    break
                            if not title:
                                title = title_elem.get_text(strip=True)
                            
                            job = {
                                'title': title,
                                'company': company_elem.get_text(strip=True) if company_elem else "N/A",
                                'link': "https://www.indeed.com" + link_elem['href'] if link_elem else "",
                                'description': description_elem.get_text(strip=True) if description_elem else "",
                                'source': 'Indeed'
                            }
                            self.jobs_data.append(job)
                    except Exception as e:
                        logger.warning(f"Error parsing Indeed job card: {str(e)}")
                        continue
                        
        except Exception as e:
            logger.error(f"Error scraping Indeed: {str(e)}")
    
    def filter_relevant_jobs(self, min_relevance_score=0.7):
        """Use Gemini to filter jobs relevant to user's resume"""
        logger.info("Filtering relevant jobs using Gemini LLM")
        relevant_jobs = []
        
        for job in self.jobs_data:
            prompt = f"""
            Evaluate how relevant this job is for a candidate with the following profile:
            
            Candidate Skills: {', '.join(self.user_skills)}
            Candidate Experience: {self.user_experience}
            
            Job Details:
            Title: {job['title']}
            Company: {job['company']}
            Description: {job['description']}
            
            On a scale of 0 to 1, how relevant is this job for the candidate?
            Also, extract the key skills required for this job as a Python list.
            
            Respond ONLY in JSON format with keys "relevance_score" (float) and "required_skills" (list).
            """
            
            try:
                response = self.model.generate_content(prompt)
                response_text = response.text.strip()
                if response_text.startswith("```json"):
                    response_text = response_text[7:]
                if response_text.endswith("```"):
                    response_text = response_text[:-3]
                    
                job_analysis = json.loads(response_text)
                relevance_score = job_analysis.get("relevance_score", 0)
                required_skills = job_analysis.get("required_skills", [])
                
                if relevance_score >= min_relevance_score:
                    job['relevance_score'] = relevance_score
                    job['required_skills'] = required_skills
                    relevant_jobs.append(job)
                    
            except Exception as e:
                logger.warning(f"Error analyzing job relevance: {str(e)}")
                # Add job with default values if analysis fails
                job['relevance_score'] = 0.5
                job['required_skills'] = ["Skills not extracted"]
                relevant_jobs.append(job)
                
        # Sort by relevance score
        relevant_jobs.sort(key=lambda x: x['relevance_score'], reverse=True)
        self.jobs_data = relevant_jobs
        logger.info(f"Filtered to {len(relevant_jobs)} relevant jobs")
        return relevant_jobs
    
    def export_to_excel(self, output_path="job_matches.xlsx"):
        """Export job matches to Excel file"""
        logger.info(f"Exporting jobs to Excel: {output_path}")
        
        # Prepare data for DataFrame
        excel_data = []
        for job in self.jobs_data:
            excel_data.append({
                'Job Title': job['title'],
                'Company': job['company'],
                'Required Skills': ', '.join(job.get('required_skills', [])),
                'Relevance Score': f"{job.get('relevance_score', 0):.2f}",
                'Apply Link': job['link'],
                'Source': job['source']
            })
        
        # Create DataFrame and export to Excel
        df = pd.DataFrame(excel_data)
        df.to_excel(output_path, index=False, engine='openpyxl')
        
        # Format Excel file
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Format header
        for col in ws.iter_cols(min_row=1, max_row=1):
            for cell in col:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        logger.info(f"Excel file saved successfully: {output_path}")
        return output_path

def main():
    # Get API key from environment variable or input
    GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
    if not GEMINI_API_KEY:
        GEMINI_API_KEY = input("Enter your Google Gemini API key: ").strip()
    
    # Initialize the agent
    agent = JobScrapingAgent(GEMINI_API_KEY)
    
    # Get user inputs
    resume_path = input("Enter path to your resume (PDF/DOCX): ").strip()
    job_title = input("Enter job title to search for: ").strip()
    location = input("Enter location: ").strip()
    
    try:
        # Extract resume information
        agent.extract_resume_info(resume_path)
        
        # Scrape jobs
        agent.scrape_jobs(job_title, location, num_pages=2)
        
        # Filter relevant jobs
        agent.filter_relevant_jobs(min_relevance_score=0.6)
        
        # Export to Excel
        output_file = f"job_matches_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        agent.export_to_excel(output_file)
        
        print(f"\nJob matching complete! Results saved to: {output_file}")
        print(f"Found {len(agent.jobs_data)} relevant job matches.")
        
    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")
        print(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    main()
